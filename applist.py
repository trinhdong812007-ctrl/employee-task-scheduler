# -*- coding: utf-8 -*-
"""
applist.py
Employee Task Scheduler - Hệ thống phân công công việc nhân viên
"""

import csv
import io
import os
import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta, date

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_, text

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__)
app.config["SECRET_KEY"] = "employee-task-scheduler-secret-key"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(BASE_DIR, "scheduler.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

CA_LAM_VIEC = {"Sáng": "07:30 - 11:30", "Chiều": "13:00 - 17:00"}
THU_TRONG_TUAN = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7"]
DO_UU_TIEN_LIST = ["Thấp", "Trung bình", "Cao", "Khẩn cấp"]
TRINH_DO_LIST = ["Cơ bản", "Khá", "Thành thạo", "Chuyên gia"]

VI_TRI_MAP = {
    "Kỹ thuật": ["Kỹ sư", "Kỹ thuật viên", "Trưởng phòng Kỹ thuật", "Phó phòng Kỹ thuật"],
    "Kinh doanh": ["Nhân viên Kinh doanh", "Trưởng phòng Kinh doanh", "Chuyên viên Kinh doanh"],
    "Kế toán": ["Kế toán viên", "Kế toán trưởng", "Kế toán tổng hợp"],
    "Hành chính": ["Nhân viên Hành chính", "Trưởng phòng Hành chính", "Thư ký"],
    "IT Support": ["Nhân viên IT", "Trưởng nhóm IT", "Chuyên viên IT"],
    "Nhân sự": ["Nhân viên Nhân sự", "Trưởng phòng Nhân sự", "Chuyên viên Nhân sự"],
    "Marketing": ["Nhân viên Marketing", "Trưởng phòng Marketing", "Chuyên viên Marketing"],
}


class Employee(db.Model):
    __tablename__ = "employee"
    id = db.Column(db.Integer, primary_key=True)
    ma_nv = db.Column(db.String(20), unique=True, nullable=False)
    ho_ten = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120))
    bo_phan = db.Column(db.String(80))
    vi_tri = db.Column(db.String(80))
    trinh_do = db.Column(db.String(20), default="Cơ bản")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    schedules = db.relationship("Schedule", backref="employee", cascade="all, delete-orphan", lazy=True)

    def to_dict(self):
        return {
            "id": self.id,
            "ma_nv": self.ma_nv,
            "ho_ten": self.ho_ten,
            "email": self.email,
            "bo_phan": self.bo_phan,
            "vi_tri": self.vi_tri,
            "trinh_do": self.trinh_do,
        }


class Task(db.Model):
    __tablename__ = "task"
    id = db.Column(db.Integer, primary_key=True)
    ma_cv = db.Column(db.String(20), unique=True, nullable=False)
    ten_cv = db.Column(db.String(150), nullable=False)
    ghi_chu = db.Column(db.Text)
    do_uu_tien = db.Column(db.String(20), default="Trung bình")
    ngay_gio = db.Column(db.DateTime)
    bo_phan = db.Column(db.String(80))
    so_luong_nv = db.Column(db.Integer, default=1)
    thoi_luong = db.Column(db.Float, default=1.0)
    ca_requirement = db.Column(db.String(20), default="Sáng")
    completed = db.Column(db.Boolean, default=False)
    completed_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    schedules = db.relationship("Schedule", backref="task", cascade="all, delete-orphan", lazy=True)

    def to_dict(self):
        return {
            "id": self.id,
            "ma_cv": self.ma_cv,
            "ten_cv": self.ten_cv,
            "ghi_chu": self.ghi_chu,
            "do_uu_tien": self.do_uu_tien,
            "ngay_gio": self.ngay_gio.strftime("%Y-%m-%d") if self.ngay_gio else "",
            "bo_phan": self.bo_phan,
            "so_luong_nv": self.so_luong_nv,
            "thoi_luong": self.thoi_luong,
        }


class Schedule(db.Model):
    __tablename__ = "schedule"
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employee.id"), nullable=False)
    task_id = db.Column(db.Integer, db.ForeignKey("task.id"), nullable=False)
    ngay_lam_viec = db.Column(db.Date, nullable=False)
    ca = db.Column(db.String(10), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    __table_args__ = (db.UniqueConstraint("employee_id", "ngay_lam_viec", "ca", name="uq_emp_day_ca"),)


class Page(db.Model):
    __tablename__ = "page"
    id = db.Column(db.Integer, primary_key=True)
    slug = db.Column(db.String(100), unique=True, nullable=False)
    tieu_de = db.Column(db.String(200), nullable=False)
    noi_dung = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


def parse_date(value, default=None):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return default or date.today()


def touch_last_update():
    update_file = os.path.join(BASE_DIR, ".last_update")
    with open(update_file, "w", encoding="utf-8") as f:
        f.write(datetime.utcnow().isoformat())
    os.utime(update_file, None)
    return update_file


def normalize_text(value):
    return str(value).strip() if value is not None else ""


def parse_date_value(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except ValueError:
        try:
            return datetime.strptime(str(value), "%d/%m/%Y").date()
        except ValueError:
            return None


def validate_employee_payload(data):
    errors = []
    ma_nv = normalize_text(data.get("ma_nv"))
    ho_ten = normalize_text(data.get("ho_ten"))
    email = normalize_text(data.get("email"))
    bo_phan = normalize_text(data.get("bo_phan"))
    vi_tri = normalize_text(data.get("vi_tri"))
    trinh_do = normalize_text(data.get("trinh_do"))

    if not ma_nv:
        errors.append("Mã nhân viên là bắt buộc.")
    if not ho_ten:
        errors.append("Họ tên là bắt buộc.")
    if not email:
        errors.append("Email là bắt buộc.")
    if not bo_phan:
        errors.append("Bộ phận là bắt buộc.")
    if not vi_tri:
        errors.append("Vị trí là bắt buộc.")
    if not trinh_do:
        errors.append("Trình độ là bắt buộc.")
    if bo_phan and bo_phan not in VI_TRI_MAP:
        errors.append("Bộ phận không hợp lệ.")
    if bo_phan and vi_tri and bo_phan in VI_TRI_MAP and vi_tri not in VI_TRI_MAP[bo_phan]:
        errors.append("Vị trí không phù hợp với bộ phận đã chọn.")
    if trinh_do and trinh_do not in TRINH_DO_LIST:
        errors.append("Trình độ không hợp lệ.")
    return errors


def validate_task_payload(data):
    errors = []
    ma_cv = normalize_text(data.get("ma_cv"))
    ten_cv = normalize_text(data.get("ten_cv"))
    do_uu_tien = normalize_text(data.get("do_uu_tien"))
    ngay_gio = normalize_text(data.get("ngay_gio"))
    bo_phan = normalize_text(data.get("bo_phan"))
    so_luong_nv = data.get("so_luong_nv")
    thoi_luong = data.get("thoi_luong")

    if not ma_cv:
        errors.append("Mã công việc là bắt buộc.")
    if not ten_cv:
        errors.append("Tên công việc là bắt buộc.")
    if not do_uu_tien:
        errors.append("Độ ưu tiên là bắt buộc.")
    if not ngay_gio:
        errors.append("Ngày công việc là bắt buộc.")
    if not bo_phan:
        errors.append("Bộ phận là bắt buộc.")
    if do_uu_tien and do_uu_tien not in DO_UU_TIEN_LIST:
        errors.append("Độ ưu tiên không hợp lệ.")
    if bo_phan and bo_phan not in VI_TRI_MAP:
        errors.append("Bộ phận không hợp lệ.")

    try:
        so_luong_nv_i = int(so_luong_nv)
        if so_luong_nv_i < 1:
            errors.append("Số lượng nhân viên phải lớn hơn 0.")
    except (TypeError, ValueError):
        errors.append("Số lượng nhân viên phải là số nguyên.")

    try:
        thoi_luong_f = float(thoi_luong)
        if thoi_luong_f <= 0:
            errors.append("Thời lượng phải lớn hơn 0.")
    except (TypeError, ValueError):
        errors.append("Thời lượng phải là số.")

    return errors


def get_week_start(d):
    return d - timedelta(days=d.weekday())


def check_trung_ca(employee_id, ngay_lam_viec, ca, exclude_id=None):
    query = Schedule.query.filter_by(employee_id=employee_id, ngay_lam_viec=ngay_lam_viec, ca=ca)
    if exclude_id:
        query = query.filter(Schedule.id != exclude_id)
    return query.first()


def get_ai_suggested_employee_ids(task, ca="Sáng", limit=None):
    if not task:
        return []
    query = Employee.query
    if task.bo_phan:
        query = query.filter(Employee.bo_phan == task.bo_phan)
    employees = query.all()
    task_date = task.ngay_gio.date() if task.ngay_gio else None
    scored = []
    for emp in employees:
        score = 0
        if emp.trinh_do == "Chuyên gia":
            score += 40
        elif emp.trinh_do == "Thành thạo":
            score += 30
        elif emp.trinh_do == "Khá":
            score += 20
        elif emp.trinh_do == "Cơ bản":
            score += 10

        available = True
        if task_date:
            existing = check_trung_ca(emp.id, task_date, ca)
            if existing:
                available = False
                score -= 100

        total_assigned = Schedule.query.filter_by(employee_id=emp.id).count()
        score -= total_assigned * 2

        scored.append({"id": emp.id, "score": score, "available": available})

    scored.sort(key=lambda x: x["score"], reverse=True)
    suggested = [item["id"] for item in scored if item["available"]]
    
    # Kiểm tra số lượng đã phân công thực tế để tránh phân công thừa
    current_assigned_count = Schedule.query.filter_by(task_id=task.id).count()
    remaining_slots = max(0, task.so_luong_nv - current_assigned_count)

    if limit:
        suggested = suggested[:min(limit, remaining_slots)]
    else:
        suggested = suggested[:remaining_slots]
    return suggested


def assign_task_ai(task, ca="Sáng"):
    if not task or not task.ngay_gio:
        return [], ["Công việc thiếu ngày hoặc không tồn tại."]
    requested_ca = ca if ca in ["Sáng", "Chiều"] else task.ca_requirement or "Sáng"
    suggested_ids = get_ai_suggested_employee_ids(task, ca=requested_ca)
    if not suggested_ids:
        return [], ["Không tìm thấy nhân viên phù hợp hoặc đã đủ số lượng."]

    assigned = []
    errors = []
    for eid in suggested_ids:
        if check_trung_ca(eid, task.ngay_gio.date(), requested_ca):
            emp = Employee.query.get(eid)
            errors.append(f"Nhân viên '{emp.ho_ten}' đã có lịch ca {requested_ca}.")
            continue
        db.session.add(Schedule(employee_id=eid, task_id=task.id, ngay_lam_viec=task.ngay_gio.date(), ca=requested_ca))
        assigned.append(eid)
    return assigned, errors


DO_UU_TIEN_COLORS = {
    "Khẩn cấp": "#ef4444",
    "Cao": "#f59e0b",
    "Trung bình": "#3b82f6",
    "Thấp": "#94a3b8",
}

DO_UU_TIEN_BG = {
    "Khẩn cấp": "rgba(239,68,68,0.18)",
    "Cao": "rgba(245,158,11,0.18)",
    "Trung bình": "rgba(59,130,246,0.18)",
    "Thấp": "rgba(148,163,184,0.18)",
}

TRINH_DO_ORDER = {"Cơ bản": 1, "Khá": 2, "Thành thạo": 3, "Chuyên gia": 4}


@app.route("/")
def index():
    return redirect(url_for("lich_trinh"))

@app.route("/lich-trinh")
def lich_trinh():
    start_param = request.args.get("start")
    if start_param:
        anchor = parse_date(start_param)
    else:
        anchor = date.today()
    week_start = get_week_start(anchor)
    week_dates = [week_start + timedelta(days=i) for i in range(6)]

    # 1. TRUY VẤN CHỈ CÁC CÔNG VIỆC ĐÃ ĐƯỢC PHÂN CÔNG TRONG BẢNG SCHEDULE
    assigned_schedules = (
        db.session.query(Schedule, Task)
        .join(Task, Schedule.task_id == Task.id)
        .filter(
            Schedule.ngay_lam_viec >= week_dates[0],
            Schedule.ngay_lam_viec <= week_dates[-1],
            Task.completed == False
        )
        .order_by(Schedule.ngay_lam_viec, Schedule.ca)
        .all()
    )

    timetable = {}
    week_tasks = []
    seen_task_ids = set()

    # 2. GOM NHÓM DỮ LIỆU ĐÃ PHÂN CÔNG
    for sch, task in assigned_schedules:
        sch_date = sch.ngay_lam_viec
        ca = sch.ca or task.ca_requirement or 'Sáng'
        key = (sch_date, ca)

        if key not in timetable:
            timetable[key] = []

        # Đảm bảo 1 công việc chỉ thêm 1 lần vào ô (dù giao cho nhiều NV)
        if task not in timetable[key]:
            timetable[key].append(task)

        # Lập danh sách công việc trong tuần (không trùng lặp)
        if task.id not in seen_task_ids:
            seen_task_ids.add(task.id)
            week_tasks.append({
                "task": task,
                "date": sch_date,
                "ca": ca
            })

    prev_week = week_start - timedelta(days=7)
    next_week = week_start + timedelta(days=7)
    total_employees = Employee.query.count()

    return render_template(
        "lich_trinh.html",
        week_start=week_start,
        week_dates=week_dates,
        thu_list=THU_TRONG_TUAN,
        timetable=timetable,
        week_tasks=week_tasks,
        prev_week=prev_week.strftime("%Y-%m-%d"),
        next_week=next_week.strftime("%Y-%m-%d"),
        today_str=date.today().strftime("%Y-%m-%d"),
        total_employees=total_employees,
        do_uu_tien_colors=DO_UU_TIEN_COLORS,
        do_uu_tien_bg=DO_UU_TIEN_BG,
    )

@app.route("/employees")
def employees_page():
    q = request.args.get("q", "").strip()
    query = Employee.query
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Employee.ma_nv.ilike(like),
                Employee.ho_ten.ilike(like),
                Employee.email.ilike(like),
                Employee.bo_phan.ilike(like),
                Employee.vi_tri.ilike(like),
                Employee.trinh_do.ilike(like),
            )
        )
    employees = query.order_by(Employee.id.desc()).all()
    return render_template("employees.html", employees=employees, q=q, vi_tri_map=VI_TRI_MAP, trinh_do_list=TRINH_DO_LIST)


@app.route("/employees/add", methods=["POST"])
def employee_add():
    payload = {
        "ma_nv": normalize_text(request.form.get("ma_nv")),
        "ho_ten": normalize_text(request.form.get("ho_ten")),
        "email": normalize_text(request.form.get("email")),
        "bo_phan": normalize_text(request.form.get("bo_phan")),
        "vi_tri": normalize_text(request.form.get("vi_tri")),
        "trinh_do": normalize_text(request.form.get("trinh_do")) or "Cơ bản",
    }
    errors = validate_employee_payload(payload)
    if errors:
        flash("; ".join(errors), "danger")
        return redirect(url_for("employees_page"))
    if Employee.query.filter_by(ma_nv=payload["ma_nv"]).first():
        flash(f"Mã nhân viên '{payload['ma_nv']}' đã tồn tại.", "danger")
        return redirect(url_for("employees_page"))
    emp = Employee(
        ma_nv=payload["ma_nv"],
        ho_ten=payload["ho_ten"],
        email=payload["email"],
        bo_phan=payload["bo_phan"],
        vi_tri=payload["vi_tri"],
        trinh_do=payload["trinh_do"],
    )
    db.session.add(emp)
    db.session.commit()
    flash(f"Đã thêm nhân viên '{payload['ho_ten']}' thành công.", "success")
    return redirect(url_for("employees_page"))


@app.route("/employees/edit/<int:emp_id>", methods=["POST"])
def employee_edit(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    payload = {
        "ma_nv": normalize_text(request.form.get("ma_nv")),
        "ho_ten": normalize_text(request.form.get("ho_ten")),
        "email": normalize_text(request.form.get("email")),
        "bo_phan": normalize_text(request.form.get("bo_phan")),
        "vi_tri": normalize_text(request.form.get("vi_tri")),
        "trinh_do": normalize_text(request.form.get("trinh_do")) or "Cơ bản",
    }
    errors = validate_employee_payload(payload)
    if errors:
        flash("; ".join(errors), "danger")
        return redirect(url_for("employees_page"))
    trung = Employee.query.filter(Employee.ma_nv == payload["ma_nv"], Employee.id != emp_id).first()
    if trung:
        flash(f"Mã nhân viên '{payload['ma_nv']}' đã được sử dụng.", "danger")
        return redirect(url_for("employees_page"))
    emp.ma_nv = payload["ma_nv"]
    emp.ho_ten = payload["ho_ten"]
    emp.email = payload["email"]
    emp.bo_phan = payload["bo_phan"]
    emp.vi_tri = payload["vi_tri"]
    emp.trinh_do = payload["trinh_do"]
    db.session.commit()
    flash("Đã cập nhật thông tin nhân viên.", "success")
    return redirect(url_for("employees_page"))


@app.route("/employees/delete/<int:emp_id>", methods=["POST"])
def employee_delete(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    db.session.delete(emp)
    db.session.commit()
    flash(f"Đã xóa nhân viên '{emp.ho_ten}'.", "success")
    return redirect(url_for("employees_page"))


@app.route("/tasks")
def tasks_page():
    q = request.args.get("q", "").strip()
    detail_id = request.args.get("detail", type=int)
    task_id = request.args.get("task_id", type=int)
    auto = request.args.get("auto") == "1"

    query = Task.query.filter_by(completed=False)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Task.ma_cv.ilike(like), Task.ten_cv.ilike(like), Task.ghi_chu.ilike(like)))
    tasks = query.order_by(Task.id.desc()).all()

    detail_task = None
    detail_assignments = []
    if detail_id:
        detail_task = Task.query.get(detail_id)
        if detail_task:
            detail_assignments = (
                db.session.query(Schedule, Employee)
                .join(Employee, Schedule.employee_id == Employee.id)
                .filter(Schedule.task_id == detail_id)
                .order_by(Schedule.ngay_lam_viec, Schedule.ca)
                .all()
            )

    suggested_task = None
    suggested_employees = []
    suggested_ids = []
    if task_id and auto:
        suggested_task = Task.query.get(task_id)
        if suggested_task:
            suggested_ids = get_ai_suggested_employee_ids(suggested_task, ca=suggested_task.ca_requirement or "Sáng")
            if suggested_ids:
                suggested_employees = Employee.query.filter(Employee.id.in_(suggested_ids)).all()
                suggested_employees.sort(key=lambda emp: suggested_ids.index(emp.id))

    return render_template(
        "tasks.html", tasks=tasks, q=q, do_uu_tien_list=DO_UU_TIEN_LIST,
        vi_tri_map=VI_TRI_MAP, trinh_do_list=TRINH_DO_LIST,
        do_uu_tien_colors=DO_UU_TIEN_COLORS, do_uu_tien_bg=DO_UU_TIEN_BG,
        detail_task=detail_task, detail_assignments=detail_assignments,
        suggested_task=suggested_task, suggested_employees=suggested_employees, suggested_ids=[str(i) for i in suggested_ids]
    )


@app.route("/tasks/add", methods=["POST"])
def task_add():
    payload = {
        "ma_cv": normalize_text(request.form.get("ma_cv")),
        "ten_cv": normalize_text(request.form.get("ten_cv")),
        "ghi_chu": normalize_text(request.form.get("ghi_chu")),
        "do_uu_tien": normalize_text(request.form.get("do_uu_tien")) or "Trung bình",
        "ngay_gio": normalize_text(request.form.get("ngay_gio")),
        "bo_phan": normalize_text(request.form.get("bo_phan")),
        "so_luong_nv": request.form.get("so_luong_nv", "1"),
        "thoi_luong": request.form.get("thoi_luong", "1"),
        "ca_requirement": normalize_text(request.form.get("ca_requirement")) or "Sáng",
    }
    errors = validate_task_payload(payload)
    if errors:
        flash("; ".join(errors), "danger")
        return redirect(url_for("tasks_page"))
    if Task.query.filter_by(ma_cv=payload["ma_cv"]).first():
        flash(f"Mã công việc '{payload['ma_cv']}' đã tồn tại.", "danger")
        return redirect(url_for("tasks_page"))
    ngay_gio = parse_date_value(payload["ngay_gio"])
    if ngay_gio is None:
        flash("Ngày công việc không đúng định dạng.", "danger")
        return redirect(url_for("tasks_page"))
    task = Task(
        ma_cv=payload["ma_cv"],
        ten_cv=payload["ten_cv"],
        ghi_chu=payload["ghi_chu"],
        do_uu_tien=payload["do_uu_tien"],
        ngay_gio=datetime.combine(ngay_gio, datetime.min.time()),
        bo_phan=payload["bo_phan"],
        so_luong_nv=int(payload["so_luong_nv"]),
        thoi_luong=float(payload["thoi_luong"]),
        ca_requirement=payload["ca_requirement"],
    )
    db.session.add(task)
    db.session.commit()
    touch_last_update()
    flash(f"Đã thêm công việc '{payload['ten_cv']}' thành công.", "success")
    return redirect(url_for("tasks_page"))


@app.route("/tasks/edit/<int:task_id>", methods=["POST"])
def task_edit(task_id):
    task = Task.query.get_or_404(task_id)
    payload = {
        "ma_cv": normalize_text(request.form.get("ma_cv")),
        "ten_cv": normalize_text(request.form.get("ten_cv")),
        "ghi_chu": normalize_text(request.form.get("ghi_chu")),
        "do_uu_tien": normalize_text(request.form.get("do_uu_tien")) or "Trung bình",
        "ngay_gio": normalize_text(request.form.get("ngay_gio")),
        "bo_phan": normalize_text(request.form.get("bo_phan")),
        "so_luong_nv": request.form.get("so_luong_nv", "1"),
        "thoi_luong": request.form.get("thoi_luong", "1"),
        "ca_requirement": normalize_text(request.form.get("ca_requirement")) or "Sáng",
    }
    errors = validate_task_payload(payload)
    if errors:
        flash("; ".join(errors), "danger")
        return redirect(url_for("tasks_page"))
    trung = Task.query.filter(Task.ma_cv == payload["ma_cv"], Task.id != task_id).first()
    if trung:
        flash(f"Mã công việc '{payload['ma_cv']}' đã được sử dụng.", "danger")
        return redirect(url_for("tasks_page"))
    task.ma_cv = payload["ma_cv"]
    task.ten_cv = payload["ten_cv"]
    task.ghi_chu = payload["ghi_chu"]
    task.do_uu_tien = payload["do_uu_tien"]
    ngay_gio = parse_date_value(payload["ngay_gio"])
    if ngay_gio is None:
        flash("Ngày công việc không đúng định dạng.", "danger")
        return redirect(url_for("tasks_page"))
    task.ngay_gio = datetime.combine(ngay_gio, datetime.min.time())
    task.bo_phan = payload["bo_phan"]
    task.so_luong_nv = int(payload["so_luong_nv"])
    task.thoi_luong = float(payload["thoi_luong"])
    old_ca = task.ca_requirement
    new_ca = payload["ca_requirement"]
    task.ca_requirement = new_ca
    task.updated_at = datetime.utcnow()

    if old_ca != new_ca:
        conflicts = []
        for sch in task.schedules:
            if sch.ca != new_ca:
                existing = Schedule.query.filter(
                    Schedule.employee_id == sch.employee_id,
                    Schedule.ngay_lam_viec == sch.ngay_lam_viec,
                    Schedule.ca == new_ca,
                    Schedule.id != sch.id,
                ).first()
                if existing:
                    employee = Employee.query.get(sch.employee_id)
                    conflicts.append(f"{employee.ho_ten} ({employee.ma_nv})")
                else:
                    sch.ca = new_ca
        if conflicts:
            flash("Không thể cập nhật ca cho một số nhân viên do trùng ca: " + ", ".join(conflicts), "warning")

    db.session.commit()
    touch_last_update()
    flash("Đã cập nhật công việc.", "success")
    return redirect(url_for("tasks_page"))


@app.route("/tasks/delete/<int:task_id>", methods=["POST"])
def task_delete(task_id):
    task = Task.query.get_or_404(task_id)
    db.session.delete(task)
    db.session.commit()
    touch_last_update()
    flash(f"Đã xóa công việc '{task.ten_cv}'.", "success")
    return redirect(url_for("tasks_page"))


@app.route("/tasks/complete/<int:task_id>", methods=["POST"])
def task_complete(task_id):
    task = Task.query.get_or_404(task_id)
    task.completed = True
    task.completed_at = datetime.utcnow()
    task.updated_at = datetime.utcnow()
    db.session.commit()
    touch_last_update()
    flash(f"Đã đánh dấu công việc '{task.ten_cv}' hoàn thành.", "success")
    return redirect(url_for("tasks_page"))


@app.route("/tasks/history")
def task_history():
    q = request.args.get("q", "").strip()
    detail_id = request.args.get("detail", type=int)
    query = Task.query.filter_by(completed=True)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Task.ma_cv.ilike(like), Task.ten_cv.ilike(like), Task.ghi_chu.ilike(like)))
    tasks = query.order_by(Task.completed_at.desc()).all()
    detail_task = None
    detail_assignments = []
    if detail_id:
        detail_task = Task.query.get(detail_id)
        if detail_task:
            detail_assignments = (
                db.session.query(Schedule, Employee)
                .join(Employee, Schedule.employee_id == Employee.id)
                .filter(Schedule.task_id == detail_id)
                .order_by(Schedule.ngay_lam_viec, Schedule.ca)
                .all()
            )
    return render_template(
        "task_history.html", tasks=tasks, q=q, do_uu_tien_list=DO_UU_TIEN_LIST,
        vi_tri_map=VI_TRI_MAP, trinh_do_list=TRINH_DO_LIST,
        do_uu_tien_colors=DO_UU_TIEN_COLORS, do_uu_tien_bg=DO_UU_TIEN_BG,
        detail_task=detail_task, detail_assignments=detail_assignments
    )


@app.route("/tasks/assign")
def task_assign():
    tasks = Task.query.filter_by(completed=False).order_by(Task.id.desc()).all()
    recent = (
        db.session.query(Schedule, Employee, Task)
        .join(Employee, Schedule.employee_id == Employee.id)
        .join(Task, Schedule.task_id == Task.id)
        .order_by(Schedule.id.desc())
        .limit(20)
        .all()
    )
    assign_mode = "auto" if request.args.get("auto") == "1" else "manual"
    return render_template(
        "task_assign.html",
        tasks=tasks,
        recent=recent,
        assign_mode=assign_mode,
        today_str=date.today().strftime("%Y-%m-%d"),
        do_uu_tien_colors=DO_UU_TIEN_COLORS,
        do_uu_tien_bg=DO_UU_TIEN_BG,
        trinh_do_list=TRINH_DO_LIST,
        vi_tri_map=VI_TRI_MAP,
    )
@app.route("/tasks/assign/delete/<int:schedule_id>", methods=["POST"])
def delete_assignment_route(schedule_id):
    sch = Schedule.query.get_or_404(schedule_id)
    task_id = sch.task_id
    
    try:
        db.session.delete(sch)
        db.session.commit()
        flash("Đã xóa phân công nhân viên thành công!", "success")
    except Exception as e:
        db.session.rollback()
        flash("Có lỗi xảy ra khi xóa!", "danger")
        
    # request.referrer giúp tự động quay về trang bạn vừa bấm nút Xóa
    return redirect(request.referrer or url_for('tasks_page', detail=task_id))

# --- ĐÃ SỬA: Lấy danh sách nhân viên chỉ phụ thuộc vào công việc ---
@app.route("/tasks/assign/get-employees", methods=["POST"])
def assign_get_employees():
    data = request.get_json()
    task_id = data.get("task_id")
    
    task = Task.query.get(task_id)
    if not task:
        return jsonify({"error": "Công việc không tồn tại"}), 404

    # Chỉ lọc nhân viên thuộc cùng bộ phận với công việc
    query = Employee.query
    if task.bo_phan:
        query = query.filter(Employee.bo_phan == task.bo_phan)

    employees = query.all()

    # Tự động lấy ca quy định của công việc (mặc định 'Sáng')
    task_ca = task.ca_requirement if task.ca_requirement in ["Sáng", "Chiều"] else "Sáng"
    task_date = task.ngay_gio.date() if task.ngay_gio else None

    result = []
    for emp in employees:
        available = True
        msg = ""
        if task_date:
            existing = check_trung_ca(emp.id, task_date, task_ca)
            if existing:
                available = False
                msg = f"Đã có lịch '{existing.task.ten_cv}' ca {task_ca}"
        result.append({
            "id": emp.id,
            "ma_nv": emp.ma_nv,
            "ho_ten": emp.ho_ten,
            "email": emp.email,
            "bo_phan": emp.bo_phan,
            "vi_tri": emp.vi_tri,
            "trinh_do": emp.trinh_do,
            "available": available,
            "msg": msg,
        })

    # Đếm số nhân viên đã được phân công thực tế
    assigned_count = Schedule.query.filter_by(task_id=task.id).count()

    return jsonify({
        "employees": result,
        "max_nv": task.so_luong_nv,
        "assigned_count": assigned_count,
        "remaining_slots": max(0, task.so_luong_nv - assigned_count),
        "ca": task_ca
    })


@app.route("/tasks/assign/ai-suggest", methods=["POST"])
def assign_ai_suggest():
    data = request.get_json()
    task_id = data.get("task_id")
    task = Task.query.get(task_id)
    if not task:
        return jsonify({"error": "Task not found"}), 404

    task_ca = task.ca_requirement if task.ca_requirement in ["Sáng", "Chiều"] else "Sáng"

    query = Employee.query
    if task.bo_phan:
        query = query.filter(Employee.bo_phan == task.bo_phan)
    employees = query.all()

    task_date = task.ngay_gio.date() if task.ngay_gio else None
    scored = []
    for emp in employees:
        score = 0
        if emp.trinh_do == "Chuyên gia":
            score += 40
        elif emp.trinh_do == "Thành thạo":
            score += 30
        elif emp.trinh_do == "Khá":
            score += 20
        elif emp.trinh_do == "Cơ bản":
            score += 10

        available = True
        if task_date:
            existing = check_trung_ca(emp.id, task_date, task_ca)
            if existing:
                available = False
                score -= 100

        total_assigned = Schedule.query.filter_by(employee_id=emp.id).count()
        score -= total_assigned * 2

        scored.append({
            "id": emp.id,
            "ma_nv": emp.ma_nv,
            "ho_ten": emp.ho_ten,
            "bo_phan": emp.bo_phan,
            "vi_tri": emp.vi_tri,
            "trinh_do": emp.trinh_do,
            "score": max(score, 0) if available else 0,
            "available": available,
            "recommended": available and score > 20,
        })

    scored.sort(key=lambda x: x["score"], reverse=True)
    return jsonify(scored[:task.so_luong_nv * 3])


# --- ĐÃ SỬA: Kiểm tra chặt chẽ giới hạn số lượng nhân viên khi lưu ---
@app.route("/tasks/assign/save", methods=["POST"])
def assign_save():
    data = request.get_json()
    task_id = data.get("task_id")
    employee_ids = data.get("employee_ids", [])
    
    task = Task.query.get(task_id)
    if not task:
        return jsonify({"error": "Công việc không tồn tại."}), 404
    if not task.ngay_gio:
        return jsonify({"error": "Công việc chưa có ngày thực hiện."}), 400

    ca = task.ca_requirement if task.ca_requirement in ["Sáng", "Chiều"] else "Sáng"

    # Đếm số lượng đã gán + số lượng đăng ký mới
    current_assigned_count = Schedule.query.filter_by(task_id=task.id).count()
    total_after_assign = current_assigned_count + len(employee_ids)

    if total_after_assign > task.so_luong_nv:
        return jsonify({
            "error": f"Vượt quá số lượng quy định! Công việc này cần tối đa {task.so_luong_nv} nhân viên (hiện tại đã phân công {current_assigned_count})."
        }), 400

    ngay = task.ngay_gio.date()
    assigned = []
    errors = []

    for eid in employee_ids:
        # Check xem nhân viên đã có trong task này chưa
        already_in_task = Schedule.query.filter_by(employee_id=eid, task_id=task_id).first()
        if already_in_task:
            emp = Employee.query.get(eid)
            errors.append(f"Nhân viên '{emp.ho_ten}' đã thuộc công việc này.")
            continue

        # Check trùng ca làm việc trong ngày
        trung = check_trung_ca(eid, ngay, ca)
        if trung:
            emp = Employee.query.get(eid)
            errors.append(f"Nhân viên '{emp.ho_ten}' đã có lịch trùng ở ca {ca}.")
            continue

        sch = Schedule(employee_id=eid, task_id=task_id, ngay_lam_viec=ngay, ca=ca)
        db.session.add(sch)
        assigned.append(eid)

    db.session.commit()
    touch_last_update()
    return jsonify({"assigned": len(assigned), "errors": errors})


@app.route("/tasks/assign/auto-all", methods=["POST"])
def auto_assign_all_tasks():
    tasks = Task.query.filter_by(completed=False).all()
    total_assigned = 0

    for task in tasks:
        if not task.ngay_gio:
            continue
        task_ca = task.ca_requirement if task.ca_requirement in ["Sáng", "Chiều"] else "Sáng"
        assigned, _ = assign_task_ai(task, ca=task_ca)
        total_assigned += len(assigned)

    db.session.commit()
    touch_last_update()
    if total_assigned:
        flash(f"AI đã tự động phân công {total_assigned} ca cho các công việc hợp lệ.", "success")
    else:
        flash("AI không thể phân công thêm ca nào. Vui lòng kiểm tra điều kiện công việc và nhân viên.", "warning")
    return redirect(url_for("tasks_page"))


@app.route('/api/last-update')
def api_last_update():
    update_file = os.path.join(BASE_DIR, ".last_update")
    if not os.path.exists(update_file):
        touch_last_update()
    mtime = datetime.utcfromtimestamp(os.path.getmtime(update_file))
    return jsonify({"last_update": mtime.isoformat()})


@app.route("/download-template/<string:template_type>")
def download_template(template_type):
    if template_type == "employees":
        filename = "employees_sample.xlsx"
    elif template_type == "tasks":
        filename = "tasks_sample.xlsx"
    else:
        return "Template không tồn tại", 404
    return send_from_directory(BASE_DIR, filename, as_attachment=True)


@app.route("/schedule/delete/<int:sch_id>", methods=["POST"])
def schedule_delete(sch_id):
    sch = Schedule.query.get_or_404(sch_id)
    db.session.delete(sch)
    db.session.commit()
    touch_last_update()
    flash("Đã xóa lịch phân công.", "success")
    return redirect(request.referrer or url_for("lich_trinh"))


@app.route("/reports")
def reports_page():
    employees = Employee.query.order_by(Employee.ho_ten).all()
    stats = []
    for emp in employees:
        count = Schedule.query.filter_by(employee_id=emp.id).count()
        total_hours = (
            db.session.query(db.func.sum(Task.thoi_luong))
            .join(Schedule, Schedule.task_id == Task.id)
            .filter(Schedule.employee_id == emp.id)
            .scalar() or 0
        )
        stats.append({"employee": emp, "so_luong_ca": count, "tong_gio": round(total_hours, 1)})
    stats.sort(key=lambda x: x["so_luong_ca"], reverse=True)
    tong_lich = Schedule.query.count()
    tong_nv_co_lich = db.session.query(Schedule.employee_id).distinct().count()
    bo_phan_stats = {}
    for emp in employees:
        bp = emp.bo_phan or "Chưa phân loại"
        bo_phan_stats[bp] = bo_phan_stats.get(bp, 0) + Schedule.query.filter_by(employee_id=emp.id).count()
    return render_template("reports.html", stats=stats, tong_lich=tong_lich, tong_nv_co_lich=tong_nv_co_lich, bo_phan_stats=bo_phan_stats)


@app.route("/api/check_conflict")
def api_check_conflict():
    employee_id = request.args.get("employee_id", type=int)
    ngay_str = request.args.get("ngay_lam_viec")
    ca = request.args.get("ca")
    if not employee_id or not ngay_str or not ca:
        return jsonify({"conflict": False})
    ngay = parse_date(ngay_str)
    trung = check_trung_ca(employee_id, ngay, ca)
    if trung:
        return jsonify({"conflict": True, "task_name": trung.task.ten_cv})
    return jsonify({"conflict": False})


@app.route("/p/<path:slug>")
def dynamic_page(slug):
    page = Page.query.filter_by(slug=slug).first()
    if not page:
        return "Trang không tồn tại (404)", 404
    return render_template("dynamic_page.html", page=page)


@app.route("/pages/add", methods=["POST"])
def page_add():
    slug = request.form.get("slug", "").strip()
    tieu_de = request.form.get("tieu_de", "").strip()
    noi_dung = request.form.get("noi_dung", "").strip()
    if not slug or not tieu_de:
        flash("Slug và Tiêu đề là bắt buộc.", "danger")
        return redirect(url_for("lich_trinh"))
    if Page.query.filter_by(slug=slug).first():
        flash(f"Đường dẫn '/p/{slug}' đã tồn tại!", "danger")
        return redirect(url_for("lich_trinh"))
    new_page = Page(slug=slug, tieu_de=tieu_de, noi_dung=noi_dung)
    db.session.add(new_page)
    db.session.commit()
    flash(f"Đã tạo đường dẫn mới: /p/{slug}", "success")
    return redirect(url_for("dynamic_page", slug=slug))


@app.route("/import-data", methods=["GET", "POST"])
def import_data_page():
    imported_count = 0
    updated_count = 0
    skipped_count = 0
    summary = []

    if request.method == "POST":
        import_type = request.form.get("import_type", "employees")
        uploaded_file = request.files.get("file")
        raw_data = request.form.get("raw_data", "").strip()

        if not uploaded_file and not raw_data:
            flash("Vui lòng chọn tập tin hoặc dán dữ liệu để nhập.", "danger")
            return redirect(url_for("import_data_page"))

        rows = []
        if uploaded_file and uploaded_file.filename:
            filename = uploaded_file.filename.lower()
            content = uploaded_file.read()
            if filename.endswith(".csv"):
                text_data = content.decode("utf-8-sig")
                rows = list(csv.DictReader(io.StringIO(text_data)))
            elif filename.endswith(".xlsx"):
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    flash("Cần cài đặt thư viện openpyxl để đọc file .xlsx.", "danger")
                    return redirect(url_for("import_data_page"))
                workbook = load_workbook(io.BytesIO(content), data_only=True)
                sheet = workbook.active
                raw_rows = list(sheet.iter_rows(values_only=True))
                if not raw_rows:
                    flash("Tập tin không chứa dữ liệu.", "danger")
                    return redirect(url_for("import_data_page"))
                headers = [normalize_text(h) for h in raw_rows[0]]
                for row in raw_rows[1:]:
                    if not any(v is not None and normalize_text(v) for v in row):
                        continue
                    record = {}
                    for index, header in enumerate(headers):
                        value = row[index] if index < len(row) else ""
                        record[header] = value
                    rows.append(record)
            else:
                flash("Chỉ hỗ trợ file .csv hoặc .xlsx.", "danger")
                return redirect(url_for("import_data_page"))
        elif raw_data:
            reader = csv.DictReader(io.StringIO(raw_data))
            if not reader.fieldnames or all(not normalize_text(fn) for fn in reader.fieldnames):
                flash("Dữ liệu CSV dán vào cần có hàng tiêu đề hợp lệ (ví dụ: ma_nv, ho_ten, bo_phan).", "danger")
                return redirect(url_for("import_data_page"))
            rows = [row for row in reader if any(normalize_text(v) for v in row.values())]

        if not rows:
            flash("Không tìm thấy dữ liệu hợp lệ trong file hoặc đoạn văn bản đã nhập.", "danger")
            return redirect(url_for("import_data_page"))

        # -----------------------------------------------------------------
        # 1. NHẬP / CẬP NHẬT NHÂN VIÊN (EMPLOYEES)
        # -----------------------------------------------------------------
        if import_type == "employees":
            for row in rows:
                payload = {
                    "ma_nv": normalize_text(row.get("ma_nv") or row.get("Mã nhân viên") or row.get("maNV") or row.get("id")),
                    "ho_ten": normalize_text(row.get("ho_ten") or row.get("Họ tên") or row.get("Họ và tên") or row.get("hoTen") or row.get("name")),
                    "email": normalize_text(row.get("email") or row.get("Email") or row.get("mail")),
                    "bo_phan": normalize_text(row.get("bo_phan") or row.get("Bộ phận") or row.get("department")),
                    "vi_tri": normalize_text(row.get("vi_tri") or row.get("Vị trí") or row.get("position")),
                    "trinh_do": normalize_text(row.get("trinh_do") or row.get("Trình độ") or row.get("level")) or "Cơ bản",
                }
                
                errors = validate_employee_payload(payload)
                if errors:
                    skipped_count += 1
                    summary.append({"row": row, "errors": errors})
                    continue

                # Tìm nhân viên đã tồn tại chưa
                existing_emp = Employee.query.filter_by(ma_nv=payload["ma_nv"]).first()

                if existing_emp:
                    # 🔄 CẬP NHẬT thông tin nhân viên đã có
                    existing_emp.ho_ten = payload["ho_ten"]
                    existing_emp.email = payload["email"]
                    existing_emp.bo_phan = payload["bo_phan"]
                    existing_emp.vi_tri = payload["vi_tri"]
                    existing_emp.trinh_do = payload["trinh_do"]
                    updated_count += 1
                else:
                    # ➕ THÊM MỚI nhân viên
                    employee = Employee(**payload)
                    db.session.add(employee)
                    imported_count += 1

            db.session.commit()
            flash(
                f"Xử lý hoàn tất! Thêm mới: {imported_count} nhân viên, Cập nhật: {updated_count} nhân viên, Bỏ qua: {skipped_count} dòng lỗi.",
                "success"
            )

        # -----------------------------------------------------------------
        # 2. NHẬP / CẬP NHẬT CÔNG VIỆC (TASKS)
        # -----------------------------------------------------------------
        else:
            for row in rows:
                payload = {
                    "ma_cv": normalize_text(row.get("ma_cv") or row.get("Mã công việc") or row.get("maCV") or row.get("id")),
                    "ten_cv": normalize_text(row.get("ten_cv") or row.get("Tên công việc") or row.get("tenCV") or row.get("title")),
                    "ghi_chu": normalize_text(row.get("ghi_chu") or row.get("Ghi chú") or row.get("note") or row.get("description")),
                    "do_uu_tien": normalize_text(row.get("do_uu_tien") or row.get("Độ ưu tiên") or row.get("priority")) or "Trung bình",
                    "ngay_gio": normalize_text(row.get("ngay_gio") or row.get("Ngày") or row.get("date") or row.get("ngay")),
                    "bo_phan": normalize_text(row.get("bo_phan") or row.get("Bộ phận") or row.get("department")),
                    "so_luong_nv": normalize_text(row.get("so_luong_nv") or row.get("Số lượng NV") or row.get("quantity") or "1"),
                    "thoi_luong": normalize_text(row.get("thoi_luong") or row.get("Thời lượng") or row.get("duration") or "1"),
                }

                errors = validate_task_payload(payload)
                if errors:
                    skipped_count += 1
                    summary.append({"row": row, "errors": errors})
                    continue

                ngay_gio = parse_date_value(payload["ngay_gio"])
                if ngay_gio is None:
                    skipped_count += 1
                    summary.append({"row": row, "errors": ["Ngày công việc không đúng định dạng."]})
                    continue

                existing_task = Task.query.filter_by(ma_cv=payload["ma_cv"]).first()

                if existing_task:
                    # 🔄 CẬP NHẬT thông tin công việc đã có
                    existing_task.ten_cv = payload["ten_cv"]
                    existing_task.ghi_chu = payload["ghi_chu"]
                    existing_task.do_uu_tien = payload["do_uu_tien"]
                    existing_task.ngay_gio = datetime.combine(ngay_gio, datetime.min.time())
                    existing_task.bo_phan = payload["bo_phan"]
                    existing_task.so_luong_nv = int(payload["so_luong_nv"])
                    existing_task.thoi_luong = float(payload["thoi_luong"])
                    updated_count += 1
                else:
                    # ➕ THÊM MỚI công việc
                    task = Task(
                        ma_cv=payload["ma_cv"],
                        ten_cv=payload["ten_cv"],
                        ghi_chu=payload["ghi_chu"],
                        do_uu_tien=payload["do_uu_tien"],
                        ngay_gio=datetime.combine(ngay_gio, datetime.min.time()),
                        bo_phan=payload["bo_phan"],
                        so_luong_nv=int(payload["so_luong_nv"]),
                        thoi_luong=float(payload["thoi_luong"]),
                    )
                    db.session.add(task)
                    imported_count += 1

            db.session.commit()
            flash(
                f"Xử lý hoàn tất! Thêm mới: {imported_count} công việc, Cập nhật: {updated_count} công việc, Bỏ qua: {skipped_count} dòng lỗi.",
                "success"
            )

        return render_template(
            "import_data.html",
            import_type=import_type,
            imported_count=imported_count,
            updated_count=updated_count,
            skipped_count=skipped_count,
            summary=summary
        )

    return render_template("import_data.html", import_type="employees", imported_count=0, updated_count=0, skipped_count=0, summary=[])


def ensure_task_schema():
    db_path = os.path.join(BASE_DIR, "scheduler.db")
    if not os.path.exists(db_path):
        return

    with db.engine.connect() as conn:
        result = conn.execute(text("PRAGMA table_info(task)"))
        columns = [row[1] for row in result]
        if "ca_requirement" not in columns:
            conn.execute(text("ALTER TABLE task ADD COLUMN ca_requirement VARCHAR(20) DEFAULT 'Sáng'"))
        if "updated_at" not in columns:
            conn.execute(text("ALTER TABLE task ADD COLUMN updated_at DATETIME"))
            conn.execute(text("UPDATE task SET updated_at = created_at WHERE updated_at IS NULL"))

        result = conn.execute(text("PRAGMA table_info(schedule)"))
        schedule_columns = [row[1] for row in result]
        if "updated_at" not in schedule_columns:
            conn.execute(text("ALTER TABLE schedule ADD COLUMN updated_at DATETIME"))
            conn.execute(text("UPDATE schedule SET updated_at = created_at WHERE updated_at IS NULL"))
        conn.execute(text("UPDATE task SET ca_requirement='Sáng' WHERE ca_requirement NOT IN ('Sáng','Chiều') OR ca_requirement IS NULL"))


with app.app_context():
    ensure_task_schema()
    db.create_all()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)